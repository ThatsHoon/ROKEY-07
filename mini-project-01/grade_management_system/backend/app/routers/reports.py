from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import Optional
from uuid import UUID
from datetime import date
import io

from app.database import get_db
from app.models.user import User
from app.models.student import Student
from app.models.course import Course, Class, Enrollment
from app.models.attendance import Attendance, AttendanceStatus
from app.models.grade import Evaluation, Grade
from app.utils.security import get_current_active_user, require_roles

router = APIRouter()


# ============ Attendance Reports ============

@router.get("/attendance/excel")
async def export_attendance_excel(
    class_id: UUID,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(["admin", "teacher", "staff"]))
):
    """출결부 Excel 출력"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    cls = db.query(Class).options(joinedload(Class.course)).filter(Class.id == str(class_id)).first()
    if not cls:
        raise HTTPException(status_code=404, detail="반을 찾을 수 없습니다")

    # Get students in class
    enrollments = db.query(Enrollment).options(
        joinedload(Enrollment.student)
    ).filter(
        Enrollment.class_id == str(class_id),
        Enrollment.dropped_at.is_(None)
    ).all()

    # Get attendance records
    query = db.query(Attendance).filter(Attendance.class_id == str(class_id))
    if date_from:
        query = query.filter(Attendance.date >= date_from)
    if date_to:
        query = query.filter(Attendance.date <= date_to)
    attendance_records = query.all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "출결부"

    # Styles
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"출결부 - {cls.course.name if cls.course else ''} {cls.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    # Get unique sessions
    sessions = sorted(set(r.session_no for r in attendance_records))

    # Header row
    headers = ['No.', '학번', '이름'] + [f'{s}회' for s in sessions] + ['출석', '지각', '결석', '출석률']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    # Data rows
    status_map = {
        AttendanceStatus.PRESENT: '○',
        AttendanceStatus.LATE: '△',
        AttendanceStatus.EARLY_LEAVE: '▽',
        AttendanceStatus.ABSENT: '×',
        AttendanceStatus.EXCUSED: '공'
    }

    for row_idx, enrollment in enumerate(enrollments, 4):
        student = enrollment.student
        if not student:
            continue

        ws.cell(row=row_idx, column=1, value=row_idx - 3).border = thin_border
        ws.cell(row=row_idx, column=2, value=student.student_number).border = thin_border
        ws.cell(row=row_idx, column=3, value=student.user.name if student.user else '').border = thin_border

        # Attendance by session
        student_attendance = {r.session_no: r for r in attendance_records if r.student_id == student.id}

        present_count = 0
        late_count = 0
        absent_count = 0

        for col_idx, session_no in enumerate(sessions, 4):
            record = student_attendance.get(session_no)
            if record:
                cell = ws.cell(row=row_idx, column=col_idx, value=status_map.get(record.status, ''))
                cell.alignment = center_align
                cell.border = thin_border

                if record.status == AttendanceStatus.PRESENT:
                    present_count += 1
                elif record.status == AttendanceStatus.LATE:
                    late_count += 1
                elif record.status == AttendanceStatus.ABSENT:
                    absent_count += 1
            else:
                ws.cell(row=row_idx, column=col_idx, value='-').border = thin_border

        # Summary columns
        col_offset = len(sessions) + 4
        ws.cell(row=row_idx, column=col_offset, value=present_count).border = thin_border
        ws.cell(row=row_idx, column=col_offset + 1, value=late_count).border = thin_border
        ws.cell(row=row_idx, column=col_offset + 2, value=absent_count).border = thin_border

        total = len(sessions)
        rate = (present_count + late_count * 0.5) / total * 100 if total > 0 else 0
        ws.cell(row=row_idx, column=col_offset + 3, value=f"{rate:.1f}%").border = thin_border

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max(max_length + 2, 8)

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    encoded_filename = quote(f"attendance_{cls.name}.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.get("/attendance/pdf")
async def export_attendance_pdf(
    class_id: UUID,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(["admin", "teacher", "staff"]))
):
    """출결부 PDF 출력"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    cls = db.query(Class).options(joinedload(Class.course)).filter(Class.id == str(class_id)).first()
    if not cls:
        raise HTTPException(status_code=404, detail="반을 찾을 수 없습니다")

    enrollments = db.query(Enrollment).options(
        joinedload(Enrollment.student)
    ).filter(
        Enrollment.class_id == str(class_id),
        Enrollment.dropped_at.is_(None)
    ).all()

    query = db.query(Attendance).filter(Attendance.class_id == str(class_id))
    if date_from:
        query = query.filter(Attendance.date >= date_from)
    if date_to:
        query = query.filter(Attendance.date <= date_to)
    attendance_records = query.all()

    # Create PDF
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=20*mm, bottomMargin=20*mm)

    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1
    )
    elements.append(Paragraph(f"Attendance Report - {cls.course.name if cls.course else ''} {cls.name}", title_style))
    elements.append(Spacer(1, 10*mm))

    # Get unique sessions
    sessions = sorted(set(r.session_no for r in attendance_records))

    # Table header
    headers = ['No.', 'ID', 'Name'] + [f'S{s}' for s in sessions] + ['P', 'L', 'A', 'Rate']

    # Table data
    status_map = {
        AttendanceStatus.PRESENT: 'O',
        AttendanceStatus.LATE: 'L',
        AttendanceStatus.EARLY_LEAVE: 'E',
        AttendanceStatus.ABSENT: 'X',
        AttendanceStatus.EXCUSED: 'EX'
    }

    table_data = [headers]

    for idx, enrollment in enumerate(enrollments, 1):
        student = enrollment.student
        if not student:
            continue

        row = [str(idx), student.student_number, student.user.name if student.user else '']

        student_attendance = {r.session_no: r for r in attendance_records if r.student_id == student.id}

        present_count = 0
        late_count = 0
        absent_count = 0

        for session_no in sessions:
            record = student_attendance.get(session_no)
            if record:
                row.append(status_map.get(record.status, '-'))
                if record.status == AttendanceStatus.PRESENT:
                    present_count += 1
                elif record.status == AttendanceStatus.LATE:
                    late_count += 1
                elif record.status == AttendanceStatus.ABSENT:
                    absent_count += 1
            else:
                row.append('-')

        total = len(sessions)
        rate = (present_count + late_count * 0.5) / total * 100 if total > 0 else 0

        row.extend([str(present_count), str(late_count), str(absent_count), f"{rate:.1f}%"])
        table_data.append(row)

    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 5*mm))
    elements.append(Paragraph("O: Present, L: Late, E: Early Leave, X: Absent, EX: Excused", styles['Normal']))

    doc.build(elements)
    output.seek(0)

    from urllib.parse import quote
    encoded_filename = quote(f"attendance_{cls.name}.pdf")
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# ============ Grade Reports ============

@router.get("/grades/excel")
async def export_grades_excel(
    course_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(["admin", "teacher", "staff"]))
):
    """성적표 Excel 출력"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    course = db.query(Course).filter(Course.id == str(course_id)).first()
    if not course:
        raise HTTPException(status_code=404, detail="과정을 찾을 수 없습니다")

    # Get evaluations
    evaluations = db.query(Evaluation).filter(Evaluation.course_id == str(course_id)).all()

    # Get students enrolled in this course
    classes = db.query(Class).filter(Class.course_id == str(course_id)).all()
    class_ids = [c.id for c in classes]

    enrollments = db.query(Enrollment).options(
        joinedload(Enrollment.student)
    ).filter(
        Enrollment.class_id.in_(class_ids),
        Enrollment.dropped_at.is_(None)
    ).all()

    # Get all grades
    grades = db.query(Grade).filter(
        Grade.evaluation_id.in_([e.id for e in evaluations])
    ).all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "성적표"

    # Styles
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"성적표 - {course.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    # Header row
    headers = ['No.', '학번', '이름']
    for eval in evaluations:
        headers.append(f"{eval.name}\n({eval.weight}%)")
    headers.extend(['총점', '등급'])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    # Data rows
    from app.schemas.grade import calculate_letter_grade
    from decimal import Decimal

    for row_idx, enrollment in enumerate(enrollments, 4):
        student = enrollment.student
        if not student:
            continue

        ws.cell(row=row_idx, column=1, value=row_idx - 3).border = thin_border
        ws.cell(row=row_idx, column=2, value=student.student_number).border = thin_border
        ws.cell(row=row_idx, column=3, value=student.user.name if student.user else '').border = thin_border

        student_grades = {g.evaluation_id: g for g in grades if g.student_id == student.id}
        total_weighted = Decimal("0")

        for col_idx, eval in enumerate(evaluations, 4):
            grade = student_grades.get(eval.id)
            if grade and grade.score is not None:
                cell = ws.cell(row=row_idx, column=col_idx, value=float(grade.score))
                weighted = (grade.score / eval.max_score) * eval.weight
                total_weighted += weighted
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value='-')
            cell.alignment = center_align
            cell.border = thin_border

        # Total and grade
        col_offset = len(evaluations) + 4
        ws.cell(row=row_idx, column=col_offset, value=float(round(total_weighted, 2))).border = thin_border
        ws.cell(row=row_idx, column=col_offset + 1, value=calculate_letter_grade(total_weighted)).border = thin_border

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max(max_length + 2, 8)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    encoded_filename = quote(f"grades_{course.code}.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.get("/grades/pdf")
async def export_grades_pdf(
    course_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(["admin", "teacher", "staff"]))
):
    """성적표 PDF 출력"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    course = db.query(Course).filter(Course.id == str(course_id)).first()
    if not course:
        raise HTTPException(status_code=404, detail="과정을 찾을 수 없습니다")

    evaluations = db.query(Evaluation).filter(Evaluation.course_id == str(course_id)).all()

    classes = db.query(Class).filter(Class.course_id == str(course_id)).all()
    class_ids = [c.id for c in classes]

    enrollments = db.query(Enrollment).options(
        joinedload(Enrollment.student)
    ).filter(
        Enrollment.class_id.in_(class_ids),
        Enrollment.dropped_at.is_(None)
    ).all()

    grades = db.query(Grade).filter(
        Grade.evaluation_id.in_([e.id for e in evaluations])
    ).all()

    # Create PDF
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=20*mm, bottomMargin=20*mm)

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1)
    elements.append(Paragraph(f"Grade Report - {course.name}", title_style))
    elements.append(Spacer(1, 10*mm))

    # Table header
    headers = ['No.', 'ID', 'Name']
    for eval in evaluations:
        headers.append(f"{eval.name}\n({eval.weight}%)")
    headers.extend(['Total', 'Grade'])

    from app.schemas.grade import calculate_letter_grade
    from decimal import Decimal

    table_data = [headers]

    for idx, enrollment in enumerate(enrollments, 1):
        student = enrollment.student
        if not student:
            continue

        row = [str(idx), student.student_number, student.user.name if student.user else '']

        student_grades = {g.evaluation_id: g for g in grades if g.student_id == student.id}
        total_weighted = Decimal("0")

        for eval in evaluations:
            grade = student_grades.get(eval.id)
            if grade and grade.score is not None:
                row.append(str(float(grade.score)))
                weighted = (grade.score / eval.max_score) * eval.weight
                total_weighted += weighted
            else:
                row.append('-')

        row.append(str(float(round(total_weighted, 2))))
        row.append(calculate_letter_grade(total_weighted))
        table_data.append(row)

    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)
    output.seek(0)

    from urllib.parse import quote
    encoded_filename = quote(f"grades_{course.code}.pdf")
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.get("/transcript/{student_id}")
async def export_student_transcript(
    student_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """개인 성적증명서 PDF 출력"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    student = db.query(Student).options(joinedload(Student.user)).filter(Student.id == str(student_id)).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # Students can only view their own transcript
    if current_user.role.name == "student" and student.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="접근 권한이 없습니다")

    # Get all grades for this student
    grades = db.query(Grade).options(
        joinedload(Grade.evaluation)
    ).filter(Grade.student_id == str(student_id)).all()

    # Create PDF
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=30*mm, bottomMargin=30*mm)

    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1)
    elements.append(Paragraph("Academic Transcript", title_style))
    elements.append(Spacer(1, 10*mm))

    # Student info
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=12)
    elements.append(Paragraph(f"Student ID: {student.student_number}", info_style))
    elements.append(Paragraph(f"Name: {student.user.name if student.user else 'Unknown'}", info_style))
    elements.append(Spacer(1, 10*mm))

    # Group grades by course
    from collections import defaultdict
    from app.schemas.grade import calculate_letter_grade
    from decimal import Decimal

    course_grades = defaultdict(list)
    for grade in grades:
        if grade.evaluation:
            course_grades[grade.evaluation.course_id].append(grade)

    for course_id, course_grade_list in course_grades.items():
        if not course_grade_list:
            continue

        course = course_grade_list[0].evaluation.course

        elements.append(Paragraph(f"Course: {course.name if course else 'Unknown'}", info_style))
        elements.append(Spacer(1, 3*mm))

        table_data = [['Evaluation', 'Score', 'Max', 'Weight', 'Weighted']]
        total_weighted = Decimal("0")

        for grade in course_grade_list:
            eval = grade.evaluation
            if grade.score is not None:
                weighted = (grade.score / eval.max_score) * eval.weight
                total_weighted += weighted
                table_data.append([
                    eval.name,
                    str(float(grade.score)),
                    str(float(eval.max_score)),
                    f"{eval.weight}%",
                    f"{float(weighted):.2f}"
                ])

        table_data.append(['Total', '', '', '', f"{float(total_weighted):.2f}"])
        table_data.append(['Grade', '', '', '', calculate_letter_grade(total_weighted)])

        table = Table(table_data, colWidths=[80*mm, 25*mm, 25*mm, 25*mm, 25*mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 10*mm))

    doc.build(elements)
    output.seek(0)

    from urllib.parse import quote
    encoded_filename = quote(f"transcript_{student.student_number}.pdf")
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
